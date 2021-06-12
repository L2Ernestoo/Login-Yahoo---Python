from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook, load_workbook
import time
driver = webdriver.Chrome("webdriver/chromedriver_91.exe")

# Credenciales desde Excel
data = load_workbook('docs/credenciales.xlsx')
de = data.active
email = de['A3'].value
password = de['B3'].value
# Url Login
driver.get('https://login.yahoo.com/')
time.sleep(3)

# ingreso Usuario y Contraseña
search_form = driver.find_element_by_id("login-username")
search_form.send_keys(email)
search_form.send_keys(Keys.ENTER)
time.sleep(3)
search_form = driver.find_element_by_id("login-passwd")
search_form.send_keys(password)
time.sleep(3)
search_form.send_keys(Keys.ENTER)

time.sleep(5)
# Redireccionamos a pantalla de Email
driver.get('https://mail.yahoo.com/?.intl=e1&.lang=es-US&.partner=none&.src=fp')
time.sleep(2)

#Click en boton de Redactar (Se Busca el boton con palabra Compose"
redactar_btn = driver.find_element_by_link_text('Compose').send_keys(Keys.ENTER)

#Datos desde Excel
datos = load_workbook('docs/datos.xlsx')

dp = datos['dp'] #Hoja Excel DP
destinatarios = datos['destinatarios'] #Hoja Excel destinatarios

#Datos en Input de Destinatario
destinatario_in = driver.find_element_by_id('message-to-field').send_keys(destinatarios['A3'].value)

#Dato en input de Asunto
asunto_in = driver.find_element_by_xpath('//input[@data-test-id="compose-subject"]').send_keys('MIS DATOS DE ESTUDIANTE - RESPUESTA DE ENUNCIADO RPA DEL EXAMEN FINAL')

#Datos en Div cuerpo de mensaje
textarea_in = driver.find_element_by_xpath('//div[@role="textbox"]').send_keys('Carne: ' + dp['B1'].value + '\n'
                                                                                'Nombre: ' + dp['B2'].value + '\n'
                                                                                'Apellido: ' + dp['B3'].value + '\n'
                                                                                'Curso: ' + dp['B4'].value + '\n'
                                                                                'Carrera: ' + dp['B5'].value + '\n'
                                                                                'Universidad: ' + dp['B6'].value + '\n'
                                                                               )

#Ejecucion en boton enviar
enviar_btn = driver.find_element_by_xpath('//button[@data-test-id="compose-send-button"]').send_keys(Keys.ENTER)

print("Mensaje Enviado...")